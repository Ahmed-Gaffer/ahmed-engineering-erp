"""
Advanced Reports API — LEGO v2 Modular
Provides multi-type reporting endpoints for the engineering module.
"""
from decimal import Decimal
from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_user
from app.projects.models import Project
from app.work_orders.models import WorkOrder
from app.engineering_features.models import (
    BOQItem, Contract, IPCHeader, DailyReport, Schedule, Subcontractor, EngDocument
)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(prefix="/api/engineering/reports", tags=["Engineering Reports"], dependencies=[Depends(get_current_user)])


# ─── Helpers ───

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
hdr_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
hdr_font = Font(bold=True, size=11, color="FFFFFF")


def style_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def style_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center' if isinstance(value, (int, float, Decimal)) else 'left')
    return cell


# ─── 1. Financial Report ───

@router.get("/financial")
async def financial_report(db: AsyncSession = Depends(get_db)):
    contracts = (await db.execute(select(Contract))).scalars().all()
    projects = (await db.execute(select(Project))).scalars().all()
    ipcs = (await db.execute(select(IPCHeader))).scalars().all()
    boq_items = (await db.execute(select(BOQItem))).scalars().all()

    total_contract_value = sum((c.value or 0) for c in contracts)
    total_boq_value = sum((b.total_price or 0) for b in boq_items)
    total_ipc_amount = sum((i.total_amount or 0) for i in ipcs)
    total_ipc_paid = sum((i.net_amount or 0) for i in ipcs if i.status == "paid")
    total_ipc_approved = sum((i.net_amount or 0) for i in ipcs if i.status == "approved")

    contract_status = {}
    for c in contracts:
        st = c.status or "unknown"
        contract_status[st] = contract_status.get(st, 0) + 1

    ipc_by_status = {}
    for i in ipcs:
        st = i.status or "unknown"
        ipc_by_status[st] = ipc_by_status.get(st, 0) + 1

    project_summary = []
    for p in projects:
        p_ipcs = [i for i in ipcs if i.project_id == p.id]
        p_contracts = [c for c in contracts if c.project_id == p.id]
        project_summary.append({
            "id": p.id,
            "code": p.code,
            "name": p.name,
            "status": p.status,
            "contract_value": float(sum((c.value or 0) for c in p_contracts)),
            "boq_value": float(sum((b.total_price or 0) for b in boq_items if b.project_id == p.id)),
            "ipc_total": float(sum((i.total_amount or 0) for i in p_ipcs)),
            "ipc_paid": float(sum((i.net_amount or 0) for i in p_ipcs if i.status == "paid")),
        })

    return {
        "total_contracts": len(contracts),
        "total_contract_value": str(total_contract_value),
        "total_boq_value": str(total_boq_value),
        "total_ipcs": len(ipcs),
        "total_ipc_amount": str(total_ipc_amount),
        "total_ipc_paid": str(total_ipc_paid),
        "total_ipc_approved": str(total_ipc_approved),
        "remaining_balance": str(total_contract_value - total_ipc_paid),
        "contracts_by_status": [{"status": k, "count": v} for k, v in contract_status.items()],
        "ipc_by_status": [{"status": k, "count": v} for k, v in ipc_by_status.items()],
        "project_breakdown": project_summary,
    }


@router.get("/financial/export")
async def export_financial_excel(db: AsyncSession = Depends(get_db)):
    data = await financial_report(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value="Financial Report — Engineering Management System").font = Font(bold=True, size=14)

    headers = ["Metric", "Value"]
    style_header(ws, headers, row=3)
    metrics = [
        ("Total Contracts", data["total_contracts"]),
        ("Total Contract Value", data["total_contract_value"]),
        ("Total BOQ Value", data["total_boq_value"]),
        ("Total IPCs", data["total_ipcs"]),
        ("Total IPC Amount", data["total_ipc_amount"]),
        ("Total IPC Paid", data["total_ipc_paid"]),
        ("Remaining Balance", data["remaining_balance"]),
    ]
    for i, (k, v) in enumerate(metrics, 4):
        style_cell(ws, i, 1, k)
        style_cell(ws, i, 2, str(v))

    proj_headers = ["Code", "Name", "Status", "Contract Value", "BOQ Value", "IPC Total", "IPC Paid"]
    style_header(ws, proj_headers, row=len(metrics) + 5)
    for i, p in enumerate(data["project_breakdown"], len(metrics) + 6):
        style_cell(ws, i, 1, p["code"])
        style_cell(ws, i, 2, p["name"])
        style_cell(ws, i, 3, p["status"])
        style_cell(ws, i, 4, p["contract_value"])
        style_cell(ws, i, 5, p["boq_value"])
        style_cell(ws, i, 6, p["ipc_total"])
        style_cell(ws, i, 7, p["ipc_paid"])

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=financial_report.xlsx"})


# ─── 2. Project Progress Report ───

@router.get("/progress")
async def progress_report(db: AsyncSession = Depends(get_db)):
    projects = (await db.execute(select(Project))).scalars().all()
    ipcs = (await db.execute(select(IPCHeader))).scalars().all()
    schedules = (await db.execute(select(Schedule))).scalars().all()

    data = []
    for p in projects:
        p_ipcs = [i for i in ipcs if i.project_id == p.id]
        p_schedules = [s for s in schedules if s.project_id == p.id]
        total_ipc_amount = sum((i.total_amount or 0) for i in p_ipcs)
        total_ipc_paid = sum((i.net_amount or 0) for i in p_ipcs if i.status == "paid")
        avg_schedule_progress = sum((s.progress_percent or 0) for s in p_schedules) / len(p_schedules) if p_schedules else 0
        has_delay = any(s.status == "delayed" for s in p_schedules)

        data.append({
            "id": p.id,
            "code": p.code,
            "name": p.name,
            "status": p.status,
            "budget_estimated": float(p.budget_estimated or 0),
            "start_date": str(p.start_date) if p.start_date else None,
            "end_date_planned": str(p.end_date_planned) if p.end_date_planned else None,
            "end_date_actual": str(p.end_date_actual) if p.end_date_actual else None,
            "ipc_total": float(total_ipc_amount),
            "ipc_paid": float(total_ipc_paid),
            "schedule_progress": float(avg_schedule_progress),
            "has_delay": has_delay,
            "schedule_count": len(p_schedules),
        })

    status_counts = {}
    for p in projects:
        st = p.status or "unknown"
        status_counts[st] = status_counts.get(st, 0) + 1

    return {
        "total_projects": len(projects),
        "projects_by_status": [{"status": k, "count": v} for k, v in status_counts.items()],
        "delayed_projects": sum(1 for d in data if d["has_delay"]),
        "avg_progress": round(sum(d["schedule_progress"] for d in data) / len(data), 2) if data else 0,
        "projects": data,
    }


@router.get("/progress/export")
async def export_progress_excel(db: AsyncSession = Depends(get_db)):
    data = await progress_report(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progress Report"
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value="Project Progress Report").font = Font(bold=True, size=14)

    headers = ["Code", "Name", "Status", "Budget", "Start", "End Planned", "End Actual",
               "IPC Total", "IPC Paid", "Schedule Progress %"]
    style_header(ws, headers, row=3)
    for i, p in enumerate(data["projects"], 4):
        style_cell(ws, i, 1, p["code"])
        style_cell(ws, i, 2, p["name"])
        style_cell(ws, i, 3, p["status"])
        style_cell(ws, i, 4, p["budget_estimated"])
        style_cell(ws, i, 5, p["start_date"] or "")
        style_cell(ws, i, 6, p["end_date_planned"] or "")
        style_cell(ws, i, 7, p["end_date_actual"] or "")
        style_cell(ws, i, 8, p["ipc_total"])
        style_cell(ws, i, 9, p["ipc_paid"])
        style_cell(ws, i, 10, p["schedule_progress"])

    for col, w in enumerate([12, 30, 15, 15, 14, 14, 14, 15, 15, 18], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=progress_report.xlsx"})


# ─── 3. Work Orders Report ───

@router.get("/work-orders")
async def work_orders_report(
    status: str = Query(None),
    priority: str = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(WorkOrder)
    if status:
        stmt = stmt.where(WorkOrder.status == status)
    if priority:
        stmt = stmt.where(WorkOrder.priority == priority)
    orders = (await db.execute(stmt.order_by(WorkOrder.id.desc()))).scalars().all()

    project_ids = list(set(o.project_id for o in orders))
    projects = {}
    if project_ids:
        projs = (await db.execute(select(Project).where(Project.id.in_(project_ids)))).scalars().all()
        projects = {p.id: p for p in projs}

    result = []
    for o in orders:
        proj = projects.get(o.project_id)
        result.append({
            "id": o.id,
            "wo_number": o.wo_number,
            "title": o.title,
            "project_id": o.project_id,
            "project_name": proj.name if proj else None,
            "project_code": proj.code if proj else None,
            "priority": o.priority,
            "status": o.status,
            "issue_date": str(o.issue_date) if o.issue_date else None,
            "due_date": str(o.due_date) if o.due_date else None,
            "total_amount": float(o.total_amount or 0),
        })

    status_counts = {}
    priority_counts = {}
    for o in orders:
        status_counts[o.status] = status_counts.get(o.status, 0) + 1
        priority_counts[o.priority] = priority_counts.get(o.priority, 0) + 1

    return {
        "total": len(orders),
        "by_status": [{"status": k, "count": v} for k, v in status_counts.items()],
        "by_priority": [{"priority": k, "count": v} for k, v in priority_counts.items()],
        "orders": result,
    }


@router.get("/work-orders/export")
async def export_work_orders_excel(db: AsyncSession = Depends(get_db)):
    data = await work_orders_report(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Work Orders"
    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="Work Orders Report").font = Font(bold=True, size=14)

    headers = ["WO #", "Title", "Project", "Priority", "Status", "Issue Date", "Due Date", "Amount"]
    style_header(ws, headers, row=3)
    for i, o in enumerate(data["orders"], 4):
        style_cell(ws, i, 1, o["wo_number"])
        style_cell(ws, i, 2, o["title"])
        style_cell(ws, i, 3, o["project_name"] or "")
        style_cell(ws, i, 4, o["priority"])
        style_cell(ws, i, 5, o["status"])
        style_cell(ws, i, 6, o["issue_date"] or "")
        style_cell(ws, i, 7, o["due_date"] or "")
        style_cell(ws, i, 8, o["total_amount"])

    for col, w in enumerate([14, 30, 25, 10, 12, 14, 14, 15], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=work_orders_report.xlsx"})


# ─── 4. Schedule Report ───

@router.get("/schedules")
async def schedule_report(
    project_id: int = Query(None),
    status: str = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Schedule)
    if project_id:
        stmt = stmt.where(Schedule.project_id == project_id)
    if status:
        stmt = stmt.where(Schedule.status == status)
    schedules = (await db.execute(stmt.order_by(Schedule.id))).scalars().all()

    project_ids = list(set(s.project_id for s in schedules))
    projects = {}
    if project_ids:
        projs = (await db.execute(select(Project).where(Project.id.in_(project_ids)))).scalars().all()
        projects = {p.id: p for p in projs}

    result = []
    for s in schedules:
        proj = projects.get(s.project_id)
        result.append({
            "id": s.id,
            "task_name": s.task_name,
            "project_id": s.project_id,
            "project_name": proj.name if proj else None,
            "project_code": proj.code if proj else None,
            "start_date": str(s.start_date) if s.start_date else None,
            "end_date": str(s.end_date) if s.end_date else None,
            "duration_days": s.duration_days,
            "progress_percent": float(s.progress_percent or 0),
            "status": s.status,
            "responsible": s.responsible,
        })

    status_counts = {}
    for s in schedules:
        status_counts[s.status] = status_counts.get(s.status, 0) + 1

    return {
        "total": len(schedules),
        "by_status": [{"status": k, "count": v} for k, v in status_counts.items()],
        "avg_progress": round(sum(s.progress_percent or 0 for s in schedules) / len(schedules), 2) if schedules else 0,
        "schedules": result,
    }


@router.get("/schedules/export")
async def export_schedules_excel(db: AsyncSession = Depends(get_db)):
    data = await schedule_report(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedules"
    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="Schedule Progress Report").font = Font(bold=True, size=14)

    headers = ["Task", "Project", "Start", "End", "Duration (Days)", "Progress %", "Status", "Responsible"]
    style_header(ws, headers, row=3)
    for i, s in enumerate(data["schedules"], 4):
        style_cell(ws, i, 1, s["task_name"])
        style_cell(ws, i, 2, s["project_name"] or "")
        style_cell(ws, i, 3, s["start_date"] or "")
        style_cell(ws, i, 4, s["end_date"] or "")
        style_cell(ws, i, 5, s["duration_days"])
        style_cell(ws, i, 6, s["progress_percent"])
        style_cell(ws, i, 7, s["status"])
        style_cell(ws, i, 8, s["responsible"] or "")

    for col, w in enumerate([30, 25, 14, 14, 16, 14, 12, 20], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=schedules_report.xlsx"})


# ─── 5. Daily Reports Summary ───

@router.get("/daily")
async def daily_reports_summary(
    project_id: int = Query(None),
    from_date: str = Query(None),
    to_date: str = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(DailyReport).order_by(DailyReport.report_date.desc())
    if project_id:
        stmt = stmt.where(DailyReport.project_id == project_id)
    if from_date:
        stmt = stmt.where(DailyReport.report_date >= date.fromisoformat(from_date))
    if to_date:
        stmt = stmt.where(DailyReport.report_date <= date.fromisoformat(to_date))

    reports = (await db.execute(stmt)).scalars().all()

    project_ids = list(set(r.project_id for r in reports))
    projects = {}
    if project_ids:
        projs = (await db.execute(select(Project).where(Project.id.in_(project_ids)))).scalars().all()
        projects = {p.id: p for p in projs}

    result = []
    for r in reports:
        proj = projects.get(r.project_id)
        result.append({
            "id": r.id,
            "project_id": r.project_id,
            "project_name": proj.name if proj else None,
            "project_code": proj.code if proj else None,
            "report_date": str(r.report_date),
            "weather": r.weather,
            "manpower_count": r.manpower_count,
            "equipment_count": r.equipment_count,
            "work_description": r.work_description,
            "issues": r.issues,
            "created_by": r.created_by,
        })

    total_manpower = sum(r.manpower_count or 0 for r in reports)
    total_equipment = sum(r.equipment_count or 0 for r in reports)

    return {
        "total_reports": len(reports),
        "total_manpower": total_manpower,
        "total_equipment": total_equipment,
        "reports": result,
    }


@router.get("/daily/export")
async def export_daily_excel(db: AsyncSession = Depends(get_db)):
    data = await daily_reports_summary(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Reports"
    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value="Daily Reports Summary").font = Font(bold=True, size=14)

    headers = ["Date", "Project", "Weather", "Manpower", "Equipment", "Work Description", "Issues", "Created By"]
    style_header(ws, headers, row=3)
    for i, r in enumerate(data["reports"], 4):
        style_cell(ws, i, 1, r["report_date"])
        style_cell(ws, i, 2, r["project_name"] or "")
        style_cell(ws, i, 3, r["weather"] or "")
        style_cell(ws, i, 4, r["manpower_count"])
        style_cell(ws, i, 5, r["equipment_count"])
        style_cell(ws, i, 6, r["work_description"] or "")
        style_cell(ws, i, 7, r["issues"] or "")
        style_cell(ws, i, 8, r["created_by"] or "")

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 20

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=daily_reports.xlsx"})
