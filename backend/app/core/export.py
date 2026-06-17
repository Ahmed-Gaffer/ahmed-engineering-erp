import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, String, or_, func
from sqlalchemy.ext.asyncio import AsyncSession


async def export_excel(
    db: AsyncSession,
    model,
    search: str | None = None,
    search_fields: list[str] | None = None,
    filters: dict | None = None,
    sort_by: str | None = None,
    sort_order: str = "asc",
    exclude_cols: set[str] = {"id", "hashed_password"},
) -> bytes:
    query = select(model)

    if filters:
        for key, value in filters.items():
            if value is not None:
                col = getattr(model, key, None)
                if col is not None:
                    query = query.where(col == value)

    if search and search_fields:
        conditions = [
            getattr(model, f).cast(String).ilike(f"%{search}%")
            for f in search_fields if hasattr(model, f)
        ]
        if conditions:
            query = query.where(or_(*conditions))

    sort_col = getattr(model, sort_by, None) if sort_by else None
    if sort_col is not None:
        query = query.order_by(sort_col.asc() if sort_order == "asc" else sort_col.desc())
    else:
        query = query.order_by(model.id)

    result = await db.execute(query)
    items = result.scalars().all()

    columns = [c for c in model.__table__.columns if c.name not in exclude_cols]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model.__tablename__

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        for col_idx, col in enumerate(columns, 1):
            val = getattr(item, col.name, "")
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col.name in ("id",) else "left")

    for col_idx, col in enumerate(columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
