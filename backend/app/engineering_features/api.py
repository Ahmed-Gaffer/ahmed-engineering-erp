import csv
from typing import List
from decimal import Decimal
from datetime import date, datetime, datetime
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, update, func
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db
from app.dependencies import get_current_user, require_role
from app.projects.models import Project

from .models import (
    BOQItem, Contract, IPCHeader, IPCDetail,
    DailyReport, Subcontractor, Schedule, EngDocument,
    VariationOrder, RFI, SystemSetting,
    MaterialApprovalRequest,
    NonConformanceReport,
    VOBOQItem, VOScheduleImpact,
)
from app.notifications.models import Notification
from .schemas import (
    ContractCreate, ContractUpdate,
    BOQItemCreate, BOQItemUpdate,
    IPCHeaderCreate, IPCHeaderUpdate,
    DrawingCreate, DailyReportCreate, DailyReportUpdate,
    SubcontractorCreate, SubcontractorUpdate,
    ScheduleCreate, ScheduleUpdate,
    VariationOrderCreate, VariationOrderUpdate,
    RFICreate, RFIUpdate,
    MARCreate, MARUpdate, MARResponse,
    NCRCreate, NCRUpdate, NCRResponse,
    NotificationCreate, NotificationResponse,
    EVMReport,
    ProjectFinancialReport, ProjectComparisonItem,
    SystemSettingCreate, SystemSettingUpdate, SystemSettingResponse,
    ActivityLogResponse,
    VOBOQItemCreate, VOBOQItemResponse,
    VOScheduleImpactCreate, VOScheduleImpactResponse,
    VOImpactSummary,
)
from app.core.audit import AuditLog
from app.auth.models import User
from .dashboard import get_dashboard_summary, get_ipc_trends
from .ipc_pdf import build_ipc_pdf
from app.workflow.engine import log_action

router = APIRouter(prefix="/api/engineering", tags=["Engineering"], dependencies=[Depends(get_current_user)])


# ─── Dashboard ───

@router.get("/dashboard/summary")
async def dashboard_summary(db: AsyncSession = Depends(get_db)):
    return await get_dashboard_summary(db)


@router.get("/dashboard/ipc-trends")
async def dashboard_ipc_trends(db: AsyncSession = Depends(get_db)):
    return await get_ipc_trends(db)


# ─── Projects ───

@router.get("/projects")
async def list_projects(skip: int = Query(0, ge=0), limit: int = Query(100, le=1000), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Project).offset(skip).limit(limit))
    return result.scalars().all()


@router.get("/projects/{project_id}")
async def get_project(project_id: int, db: AsyncSession = Depends(get_db)):
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    return project


# ─── Contracts ───

@router.post("/contracts", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_contract(data: ContractCreate, db: AsyncSession = Depends(get_db)):
    contract = Contract(**data.model_dump())
    db.add(contract)
    await db.commit()
    await db.refresh(contract)
    return contract


@router.get("/projects/{project_id}/contracts")
async def list_project_contracts(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Contract).where(Contract.project_id == project_id))
    return result.scalars().all()


@router.get("/contracts/{contract_id}")
async def get_contract(contract_id: int, db: AsyncSession = Depends(get_db)):
    contract = await db.get(Contract, contract_id)
    if not contract:
        raise HTTPException(404, "Contract not found")
    return contract


@router.patch("/contracts/{contract_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_contract(contract_id: int, data: ContractUpdate, db: AsyncSession = Depends(get_db)):
    contract = await db.get(Contract, contract_id)
    if not contract:
        raise HTTPException(404, "Contract not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(contract, key, val)
    await db.commit()
    await db.refresh(contract)
    return contract


@router.delete("/contracts/{contract_id}", dependencies=[Depends(require_role("admin"))])
async def delete_contract(contract_id: int, db: AsyncSession = Depends(get_db)):
    contract = await db.get(Contract, contract_id)
    if not contract:
        raise HTTPException(404, "Contract not found")
    await db.delete(contract)
    await db.commit()
    return {"detail": "Contract deleted successfully"}


# ─── BOQ ───

@router.post("/boq-items", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_boq_item(data: BOQItemCreate, db: AsyncSession = Depends(get_db)):
    item_data = data.model_dump()
    item_data["total_price"] = data.quantity * data.unit_price
    item = BOQItem(**item_data)
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return item


@router.post("/boq-items/bulk", dependencies=[Depends(require_role("admin", "engineer"))])
async def bulk_create_boq_items(items: List[BOQItemCreate], db: AsyncSession = Depends(get_db)):
    created = []
    for data in items:
        item_data = data.model_dump()
        item_data["total_price"] = data.quantity * data.unit_price
        item = BOQItem(**item_data)
        db.add(item)
        created.append(item)
    await db.commit()
    for item in created:
        await db.refresh(item)
    return created


@router.get("/projects/{project_id}/boq")
async def list_project_boq(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(BOQItem).where(BOQItem.project_id == project_id))
    items = result.scalars().all()
    return items


@router.put("/boq-items/{item_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_boq_item(item_id: int, data: BOQItemUpdate, db: AsyncSession = Depends(get_db)):
    item = await db.get(BOQItem, item_id)
    if not item:
        raise HTTPException(404, "BOQ item not found")
    update_data = data.model_dump(exclude_unset=True)
    for key, val in update_data.items():
        setattr(item, key, val)
    if "quantity" in update_data or "unit_price" in update_data:
        item.total_price = item.quantity * item.unit_price
    await db.commit()
    await db.refresh(item)
    return item


@router.delete("/boq-items/{item_id}", dependencies=[Depends(require_role("admin"))])
async def delete_boq_item(item_id: int, db: AsyncSession = Depends(get_db)):
    item = await db.get(BOQItem, item_id)
    if not item:
        raise HTTPException(404, "BOQ item not found")
    await db.delete(item)
    await db.commit()
    return {"detail": "BOQ item deleted successfully"}


# ─── BOQ Export/Import ───

@router.get("/projects/{project_id}/boq/export")
async def export_boq_excel(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(BOQItem).where(BOQItem.project_id == project_id))
    items = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["Item Code", "Description", "Unit", "Quantity", "Unit Price", "Total Price", "Category", "Group"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item.item_code).border = thin_border
        ws.cell(row=i, column=2, value=item.description).border = thin_border
        ws.cell(row=i, column=3, value=item.unit).border = thin_border
        ws.cell(row=i, column=4, value=float(item.quantity)).border = thin_border
        ws.cell(row=i, column=5, value=float(item.unit_price)).border = thin_border
        ws.cell(row=i, column=6, value=float(item.total_price)).border = thin_border
        ws.cell(row=i, column=7, value=item.category or "").border = thin_border
        ws.cell(row=i, column=8, value="Yes" if item.is_group else "No").border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 8

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=boq_project_{project_id}.xlsx"})


@router.post("/projects/{project_id}/boq/import", dependencies=[Depends(require_role("admin", "engineer"))])
async def import_boq_excel(project_id: int, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    errors = []
    for i, row in enumerate(rows):
        try:
            item_code, description, unit, quantity, unit_price, _, category, is_group = row
            is_group_val = str(is_group).lower() in ("yes", "true", "1") if is_group else False
            item = BOQItem(
                project_id=project_id,
                item_code=str(item_code or ""),
                description=str(description or ""),
                unit=str(unit or ""),
                quantity=Decimal(str(quantity or 0)),
                unit_price=Decimal(str(unit_price or 0)),
                category=str(category or "") if category else None,
                is_group=is_group_val,
            )
            item.total_price = item.quantity * item.unit_price
            db.add(item)
            created += 1
        except Exception as e:
            errors.append({"row": i + 2, "error": str(e)})
    await db.commit()
    return {"created": created, "errors": errors}


# ─── IPC ───

@router.post("/ipcs", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_ipc(data: IPCHeaderCreate, db: AsyncSession = Depends(get_db)):
    contract = await db.get(Contract, data.contract_id) if data.contract_id else None
    contract_value = contract.value if contract else Decimal("0")
    ret_percent = data.model_dump().get("retention_percent", contract.retention_percent if contract else Decimal("0"))

    header = IPCHeader(
        project_id=data.project_id,
        contract_id=data.contract_id,
        ipc_number=data.ipc_number,
        ipc_period=data.ipc_period,
        start_date=data.start_date,
        end_date=data.end_date,
        advance_recovery=data.advance_recovery,
        materials_on_site=data.materials_on_site,
        fines=data.fines,
        insurance=data.insurance,
        other_deductions=data.other_deductions,
        retention_percent=ret_percent,
        contract_ceiling=contract_value,
    )
    db.add(header)
    await db.flush()

    total_works = Decimal("0")
    for detail_data in data.details:
        boq = await db.get(BOQItem, detail_data.boq_item_id)
        if not boq:
            continue

        prev_result = await db.execute(
            select(IPCDetail).where(IPCDetail.boq_item_id == detail_data.boq_item_id).order_by(IPCDetail.id.desc())
        )
        prev = prev_result.scalars().all()
        previous_qty = sum(d.current_quantity for d in prev)
        current_qty = detail_data.current_quantity
        cumulative = previous_qty + current_qty
        amount = current_qty * boq.unit_price
        total_works += amount

        detail = IPCDetail(
            ipc_id=header.id,
            boq_item_id=detail_data.boq_item_id,
            previous_quantity=previous_qty,
            current_quantity=current_qty,
            cumulative_quantity=cumulative,
            percentage=(cumulative / boq.quantity * 100) if boq.quantity > 0 else Decimal("0"),
            amount=amount,
            boq_item_code=boq.item_code,
            boq_item_description=boq.description,
            boq_item_unit=boq.unit,
        )
        db.add(detail)

    gross = total_works + data.materials_on_site
    retention_amt = gross * (ret_percent / 100)
    total_ded = retention_amt + data.advance_recovery + data.fines + data.insurance + data.other_deductions
    net = gross - total_ded

    prev_ipcs = await db.execute(
        select(IPCHeader).where(
            IPCHeader.project_id == data.project_id,
            IPCHeader.id != header.id
        )
    )
    all_prev = prev_ipcs.scalars().all()
    previous_total = sum(p.net_amount for p in all_prev)
    total_billed = previous_total + net

    header.total_works = total_works
    header.gross_amount = gross
    header.retention_amount = retention_amt
    header.total_deductions = total_ded
    header.net_amount = net
    header.previous_total = previous_total
    header.current_due = net
    header.total_to_date = total_billed
    header.total_billed = total_billed
    await db.commit()
    await db.refresh(header)
    return header


@router.get("/projects/{project_id}/ipcs")
async def list_project_ipcs(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(IPCHeader).where(IPCHeader.project_id == project_id))
    return result.scalars().all()


@router.get("/ipcs/{ipc_id}")
async def get_ipc(ipc_id: int, db: AsyncSession = Depends(get_db)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc:
        raise HTTPException(404, "IPC not found")
    return ipc


@router.post("/ipcs/{ipc_id}/submit", dependencies=[Depends(require_role("admin", "engineer"))])
async def submit_ipc(ipc_id: int, comment: str = None, assigned_to: str = None, db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc or ipc.status != "draft":
        raise HTTPException(400, "IPC cannot be submitted")
    old_status = ipc.status
    ipc.status = "submitted"
    await _create_notification(db, user.id, f"IPC {ipc.ipc_number} Submitted", f"IPC {ipc.ipc_number} has been submitted for approval.", "info", "/engineering/ipc")
    await db.commit()
    await log_action(db, "ipc", ipc_id, "submit", old_status, "submitted", user.id, getattr(user, 'name', None), assigned_to, comment)
    await db.refresh(ipc)
    return ipc


@router.post("/ipcs/{ipc_id}/approve", dependencies=[Depends(require_role("admin", "engineer"))])
async def approve_ipc(ipc_id: int, comment: str = None, assigned_to: str = None, db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc or ipc.status not in ("draft", "submitted"):
        raise HTTPException(400, "IPC cannot be approved")
    old_status = ipc.status
    ipc.status = "approved"
    await _create_notification(db, user.id, f"IPC {ipc.ipc_number} Approved", f"IPC {ipc.ipc_number} has been approved.", "success", "/engineering/ipc")
    await db.commit()
    await log_action(db, "ipc", ipc_id, "approve", old_status, "approved", user.id, getattr(user, 'name', None), assigned_to, comment)
    await db.refresh(ipc)
    return ipc


@router.post("/ipcs/{ipc_id}/reject", dependencies=[Depends(require_role("admin", "engineer"))])
async def reject_ipc(ipc_id: int, comment: str = None, assigned_to: str = None, db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc or ipc.status != "submitted":
        raise HTTPException(400, "IPC cannot be rejected")
    old_status = ipc.status
    ipc.status = "rejected"
    await _create_notification(db, user.id, f"IPC {ipc.ipc_number} Rejected", f"IPC {ipc.ipc_number} has been rejected.", "error", "/engineering/ipc")
    await db.commit()
    await log_action(db, "ipc", ipc_id, "reject", old_status, "rejected", user.id, getattr(user, 'name', None), assigned_to, comment)
    await db.refresh(ipc)
    return ipc


@router.post("/ipcs/{ipc_id}/pay", dependencies=[Depends(require_role("admin", "engineer"))])
async def pay_ipc(ipc_id: int, comment: str = None, assigned_to: str = None, db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc or ipc.status != "approved":
        raise HTTPException(400, "IPC cannot be paid")
    old_status = ipc.status
    ipc.status = "paid"
    await _create_notification(db, user.id, f"IPC {ipc.ipc_number} Paid", f"IPC {ipc.ipc_number} has been marked as paid.", "success", "/engineering/ipc")
    await db.commit()
    await log_action(db, "ipc", ipc_id, "pay", old_status, "paid", user.id, getattr(user, 'name', None), assigned_to, comment)
    await db.refresh(ipc)
    return ipc


@router.delete("/ipcs/{ipc_id}", dependencies=[Depends(require_role("admin"))])
async def delete_ipc(ipc_id: int, db: AsyncSession = Depends(get_db)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc:
        raise HTTPException(404, "IPC not found")
    await db.delete(ipc)
    await db.commit()
    return {"detail": "IPC deleted successfully"}


@router.put("/ipcs/{ipc_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_ipc(ipc_id: int, data: IPCHeaderUpdate, db: AsyncSession = Depends(get_db)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc:
        raise HTTPException(404, "IPC not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(ipc, key, val)
    recalc_fields = {"materials_on_site", "fines", "insurance", "other_deductions"}
    if recalc_fields & set(data.model_dump(exclude_unset=True).keys()):
        gross = ipc.total_works + ipc.materials_on_site
        retention_amt = gross * (ipc.retention_percent / 100) if ipc.retention_percent else Decimal("0")
        ipc.gross_amount = gross
        ipc.retention_amount = retention_amt
        ipc.total_deductions = retention_amt + ipc.advance_recovery + ipc.fines + ipc.insurance + ipc.other_deductions
        ipc.net_amount = gross - ipc.total_deductions
        ipc.current_due = ipc.net_amount
        ipc.total_to_date = ipc.previous_total + ipc.net_amount
        ipc.total_billed = ipc.total_to_date
    await db.commit()
    await db.refresh(ipc)
    return ipc


# ─── IPC Export ───

@router.get("/ipcs/{ipc_id}/export")
async def export_ipc_excel(ipc_id: int, db: AsyncSession = Depends(get_db)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc:
        raise HTTPException(404, "IPC not found")
    result = await db.execute(select(IPCDetail).where(IPCDetail.ipc_id == ipc_id))
    details = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"IPC-{ipc.ipc_number}"
    hdr_font = Font(bold=True, size=12)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.cell(row=1, column=1, value=f"IPC: {ipc.ipc_number}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Period: {ipc.start_date} to {ipc.end_date}")
    ws.cell(row=3, column=1, value=f"Status: {ipc.status}")
    ws.cell(row=4, column=1, value=f"Total Works: {ipc.total_works} | Gross: {ipc.gross_amount} | Net Due: {ipc.net_amount} | Retention: {ipc.retention_amount}")
    ws.cell(row=5, column=1, value=f"Deductions - Fines: {ipc.fines} | Insurance: {ipc.insurance} | Other: {ipc.other_deductions} | Total: {ipc.total_deductions}")
    ws.cell(row=6, column=1, value=f"Cumulative - Previous: {ipc.previous_total} | Current Due: {ipc.current_due} | Total to Date: {ipc.total_to_date} | Ceiling: {ipc.contract_ceiling} | Billed: {ipc.total_billed}")

    headers = ["Item Code", "Description", "Unit", "Previous Qty", "Current Qty", "Cumulative", "Amount"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=h)
        cell.font = hdr_font_white
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, d in enumerate(details, 9):
        ws.cell(row=i, column=1, value=d.boq_item_code).border = thin_border
        ws.cell(row=i, column=2, value=d.boq_item_description).border = thin_border
        ws.cell(row=i, column=3, value=d.boq_item_unit).border = thin_border
        ws.cell(row=i, column=4, value=float(d.previous_quantity)).border = thin_border
        ws.cell(row=i, column=5, value=float(d.current_quantity)).border = thin_border
        ws.cell(row=i, column=6, value=float(d.cumulative_quantity)).border = thin_border
        ws.cell(row=i, column=7, value=float(d.amount)).border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=ipc_{ipc.ipc_number}.xlsx"})


# ─── IPC PDF ───

@router.get("/ipcs/{ipc_id}/pdf")
async def export_ipc_pdf(ipc_id: int, db: AsyncSession = Depends(get_db)):
    ipc = await db.get(IPCHeader, ipc_id)
    if not ipc:
        raise HTTPException(404, "IPC not found")
    result = await db.execute(select(IPCDetail).where(IPCDetail.ipc_id == ipc_id))
    details = result.scalars().all()
    project = await db.get(Project, ipc.project_id)
    pdf_buf = build_ipc_pdf(ipc, details, project.name if project else "")
    return StreamingResponse(
        pdf_buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=ipc_{ipc.ipc_number}.pdf"},
    )


# ─── IPC Project Summary ───

@router.get("/projects/{project_id}/ipcs/summary")
async def ipc_project_summary(project_id: int, db: AsyncSession = Depends(get_db)):
    """Aggregated IPC summary for a project with cumulative financial data."""
    from decimal import Decimal

    result = await db.execute(
        select(IPCHeader)
        .where(IPCHeader.project_id == project_id)
        .order_by(IPCHeader.ipc_period.asc(), IPCHeader.id.asc())
    )
    ipcs = result.scalars().all()

    if not ipcs:
        return {
            "project_id": project_id,
            "total_ipcs": 0,
            "total_net_amount": "0",
            "total_works": "0",
            "total_paid": "0",
            "total_pending": "0",
            "contract_ceiling": "0",
            "total_billed": "0",
            "remaining": "0",
            "execution_percent": "0",
            "ipcs": [],
            "by_status": [],
        }

    latest = ipcs[-1]
    total_paid = sum((i.net_amount or 0) for i in ipcs if i.status == "paid")
    total_pending = sum((i.net_amount or 0) for i in ipcs if i.status in ("draft", "submitted"))

    by_status = {}
    for ipc in ipcs:
        st = ipc.status or "unknown"
        if st not in by_status:
            by_status[st] = {"count": 0, "amount": Decimal("0")}
        by_status[st]["count"] += 1
        by_status[st]["amount"] += ipc.net_amount or 0

    return {
        "project_id": project_id,
        "total_ipcs": len(ipcs),
        "total_net_amount": str(latest.total_to_date),
        "total_works": str(latest.total_billed or 0),
        "total_paid": str(total_paid),
        "total_pending": str(total_pending),
        "contract_ceiling": str(latest.contract_ceiling),
        "total_billed": str(latest.total_billed),
        "remaining": str(Decimal(str(latest.contract_ceiling or 0)) - Decimal(str(latest.total_billed or 0))),
        "execution_percent": str(round((latest.total_billed / latest.contract_ceiling * 100) if latest.contract_ceiling and latest.contract_ceiling > 0 else 0, 2)),
        "ipcs": [
            {
                "id": ipc.id,
                "ipc_number": ipc.ipc_number,
                "ipc_period": ipc.ipc_period,
                "status": ipc.status,
                "net_amount": str(ipc.net_amount),
                "total_works": str(ipc.total_works),
                "total_to_date": str(ipc.total_to_date),
                "current_due": str(ipc.current_due),
                "start_date": str(ipc.start_date) if ipc.start_date else None,
                "end_date": str(ipc.end_date) if ipc.end_date else None,
                "created_at": str(ipc.created_at) if ipc.created_at else None,
            }
            for ipc in ipcs
        ],
        "by_status": [
            {"status": k, "count": v["count"], "amount": str(v["amount"])}
            for k, v in by_status.items()
        ],
    }


@router.get("/ipcs/{ipc_id}/comparison")
async def ipc_comparison(ipc_id: int, db: AsyncSession = Depends(get_db)):
    """Compare an IPC with the previous IPC for the same project+contract."""
    from decimal import Decimal
    from sqlalchemy import select

    current = await db.get(IPCHeader, ipc_id)
    if not current:
        raise HTTPException(404, "IPC not found")

    prev_result = await db.execute(
        select(IPCHeader)
        .where(
            IPCHeader.project_id == current.project_id,
            IPCHeader.contract_id == current.contract_id,
            IPCHeader.id < current.id,
        )
        .order_by(IPCHeader.id.desc())
        .limit(1)
    )
    previous = prev_result.scalar_one_or_none()

    def field_diff(name, current_val, previous_val):
        curr = Decimal(str(current_val or 0))
        prev = Decimal(str(previous_val or 0))
        diff = curr - prev
        pct = (diff / prev * 100) if prev != 0 else None
        return {
            "field": name,
            "current": str(curr),
            "previous": str(prev) if previous else "0",
            "variance": str(diff),
            "variance_percent": str(round(pct, 2)) if pct is not None else None,
        }

    financial_fields = [
        ("total_works", current.total_works, previous.total_works if previous else 0),
        ("materials_on_site", current.materials_on_site, previous.materials_on_site if previous else 0),
        ("gross_amount", current.gross_amount, previous.gross_amount if previous else 0),
        ("retention_amount", current.retention_amount, previous.retention_amount if previous else 0),
        ("advance_recovery", current.advance_recovery, previous.advance_recovery if previous else 0),
        ("fines", current.fines, previous.fines if previous else 0),
        ("insurance", current.insurance, previous.insurance if previous else 0),
        ("other_deductions", current.other_deductions, previous.other_deductions if previous else 0),
        ("total_deductions", current.total_deductions, previous.total_deductions if previous else 0),
        ("net_amount", current.net_amount, previous.net_amount if previous else 0),
        ("total_to_date", current.total_to_date, previous.total_to_date if previous else 0),
    ]

    return {
        "current": {
            "id": current.id,
            "ipc_number": current.ipc_number,
            "ipc_period": current.ipc_period,
            "status": current.status,
            "start_date": str(current.start_date) if current.start_date else None,
            "end_date": str(current.end_date) if current.end_date else None,
            "total_works": str(current.total_works),
            "gross_amount": str(current.gross_amount),
            "total_deductions": str(current.total_deductions),
            "net_amount": str(current.net_amount),
            "total_to_date": str(current.total_to_date),
            "contract_ceiling": str(current.contract_ceiling),
            "total_billed": str(current.total_billed),
            "remaining": str(Decimal(str(current.contract_ceiling or 0)) - Decimal(str(current.total_billed or 0))),
        },
        "previous": {
            "id": previous.id,
            "ipc_number": previous.ipc_number,
            "net_amount": str(previous.net_amount),
            "total_to_date": str(previous.total_to_date),
        } if previous else None,
        "variance_analysis": [field_diff(*f) for f in financial_fields],
        "cumulative": {
            "total_ipcs": current.ipc_period,
            "total_net_amount": str(current.total_to_date),
            "contract_ceiling": str(current.contract_ceiling),
            "total_billed": str(current.total_billed),
            "remaining": str(Decimal(str(current.contract_ceiling or 0)) - Decimal(str(current.total_billed or 0))),
            "execution_percent": str(round((current.total_billed / current.contract_ceiling * 100) if current.contract_ceiling and current.contract_ceiling > 0 else 0, 2)),
        },
    }


# ─── Drawings ───


@router.post("/drawings", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_drawing(data: DrawingCreate, db: AsyncSession = Depends(get_db)):
    from app.drawings.models import Drawing
    dwg = Drawing(**data.model_dump())
    db.add(dwg)
    await db.commit()
    await db.refresh(dwg)
    return dwg

@router.get("/projects/{project_id}/drawings")
async def list_project_drawings(project_id: int, db: AsyncSession = Depends(get_db)):
    from app.drawings.models import Drawing
    result = await db.execute(select(Drawing).where(Drawing.project_id == project_id))
    return result.scalars().all()


# ─── Daily Reports ───

@router.post("/daily-reports", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_daily_report(data: DailyReportCreate, db: AsyncSession = Depends(get_db)):
    report = DailyReport(**data.model_dump())
    db.add(report)
    await db.commit()
    await db.refresh(report)
    return report


@router.get("/projects/{project_id}/daily-reports")
async def list_daily_reports(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DailyReport).where(DailyReport.project_id == project_id))
    return result.scalars().all()


@router.get("/daily-reports/{report_id}")
async def get_daily_report(report_id: int, db: AsyncSession = Depends(get_db)):
    report = await db.get(DailyReport, report_id)
    if not report:
        raise HTTPException(404, "Daily report not found")
    return report


@router.put("/daily-reports/{report_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_daily_report(report_id: int, data: DailyReportUpdate, db: AsyncSession = Depends(get_db)):
    report = await db.get(DailyReport, report_id)
    if not report:
        raise HTTPException(404, "Daily report not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(report, key, val)
    await db.commit()
    await db.refresh(report)
    return report


@router.delete("/daily-reports/{report_id}", dependencies=[Depends(require_role("admin"))])
async def delete_daily_report(report_id: int, db: AsyncSession = Depends(get_db)):
    report = await db.get(DailyReport, report_id)
    if not report:
        raise HTTPException(404, "Daily report not found")
    await db.delete(report)
    await db.commit()
    return {"detail": "Daily report deleted successfully"}


# ─── Subcontractors ───

@router.post("/subcontractors", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_subcontractor(data: SubcontractorCreate, db: AsyncSession = Depends(get_db)):
    sub = Subcontractor(**data.model_dump())
    db.add(sub)
    await db.commit()
    await db.refresh(sub)
    return sub


@router.get("/projects/{project_id}/subcontractors")
async def list_subcontractors(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Subcontractor).where(Subcontractor.project_id == project_id))
    return result.scalars().all()


@router.get("/subcontractors/{sub_id}")
async def get_subcontractor(sub_id: int, db: AsyncSession = Depends(get_db)):
    sub = await db.get(Subcontractor, sub_id)
    if not sub:
        raise HTTPException(404, "Subcontractor not found")
    return sub


@router.put("/subcontractors/{sub_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_subcontractor(sub_id: int, data: SubcontractorUpdate, db: AsyncSession = Depends(get_db)):
    sub = await db.get(Subcontractor, sub_id)
    if not sub:
        raise HTTPException(404, "Subcontractor not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(sub, key, val)
    await db.commit()
    await db.refresh(sub)
    return sub


@router.delete("/subcontractors/{sub_id}", dependencies=[Depends(require_role("admin"))])
async def delete_subcontractor(sub_id: int, db: AsyncSession = Depends(get_db)):
    sub = await db.get(Subcontractor, sub_id)
    if not sub:
        raise HTTPException(404, "Subcontractor not found")
    await db.delete(sub)
    await db.commit()
    return {"detail": "Subcontractor deleted successfully"}


# ─── Schedules / Gantt ───

@router.post("/schedules", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_schedule(data: ScheduleCreate, db: AsyncSession = Depends(get_db)):
    sched = Schedule(**data.model_dump())
    db.add(sched)
    await db.commit()
    await db.refresh(sched)
    return sched


@router.get("/projects/{project_id}/schedules")
async def list_schedules(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Schedule).where(Schedule.project_id == project_id))
    return result.scalars().all()


@router.patch("/schedules/{schedule_id}/progress", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_schedule_progress(schedule_id: int, progress: float, db: AsyncSession = Depends(get_db)):
    sched = await db.get(Schedule, schedule_id)
    if not sched:
        raise HTTPException(404, "Schedule not found")
    sched.progress_percent = Decimal(str(progress))
    await db.commit()
    await db.refresh(sched)
    return sched


@router.get("/schedules/{schedule_id}")
async def get_schedule(schedule_id: int, db: AsyncSession = Depends(get_db)):
    sched = await db.get(Schedule, schedule_id)
    if not sched:
        raise HTTPException(404, "Schedule not found")
    return sched


@router.put("/schedules/{schedule_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_schedule(schedule_id: int, data: ScheduleUpdate, db: AsyncSession = Depends(get_db)):
    sched = await db.get(Schedule, schedule_id)
    if not sched:
        raise HTTPException(404, "Schedule not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(sched, key, val)
    await db.commit()
    await db.refresh(sched)
    return sched


@router.delete("/schedules/{schedule_id}", dependencies=[Depends(require_role("admin"))])
async def delete_schedule(schedule_id: int, db: AsyncSession = Depends(get_db)):
    sched = await db.get(Schedule, schedule_id)
    if not sched:
        raise HTTPException(404, "Schedule not found")
    await db.delete(sched)
    await db.commit()
    return {"detail": "Schedule deleted successfully"}


@router.post("/projects/{project_id}/schedules/critical-path")
async def calculate_critical_path(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Schedule).where(Schedule.project_id == project_id))
    tasks = result.scalars().all()
    task_map = {t.id: t for t in tasks}

    for task in tasks:
        task.critical = False

    for task in tasks:
        deps = [int(d.strip()) for d in (task.dependencies or "").split(",") if d.strip().isdigit()]
        if not deps:
            task.critical = True

    critical_ids = set()
    for task in tasks:
        deps = [int(d.strip()) for d in (task.dependencies or "").split(",") if d.strip().isdigit()]
        if not deps:
            critical_ids.add(task.id)
        for dep_id in deps:
            if dep_id in critical_ids:
                critical_ids.add(task.id)

    for task in tasks:
        task.critical = task.id in critical_ids

    await db.commit()
    return {"critical_count": len(critical_ids), "total": len(tasks)}


@router.patch("/schedules/{schedule_id}/dependencies")
async def update_dependencies(schedule_id: int, dependencies: str, db: AsyncSession = Depends(get_db)):
    sched = await db.get(Schedule, schedule_id)
    if not sched:
        raise HTTPException(404, "Schedule not found")
    sched.dependencies = dependencies
    await db.commit()
    await db.refresh(sched)
    return sched


# ─── Documents ───

@router.post("/documents", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_document(
    project_id: int, title: str, doc_type: str = "correspondence",
    reference_number: str = None, file_path: str = None,
    status: str = "draft", created_by: str = None,
    db: AsyncSession = Depends(get_db),
):
    doc = EngDocument(
        project_id=project_id, title=title, doc_type=doc_type,
        reference_number=reference_number, file_path=file_path,
        status=status, created_by=created_by,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    return doc


@router.get("/projects/{project_id}/documents")
async def list_documents(project_id: int, doc_type: str = Query(None), db: AsyncSession = Depends(get_db)):
    stmt = select(EngDocument).where(EngDocument.project_id == project_id)
    if doc_type:
        stmt = stmt.where(EngDocument.doc_type == doc_type)
    result = await db.execute(stmt)
    return result.scalars().all()


# ─── Variation Orders ───

@router.post("/variation-orders", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_vo(data: VariationOrderCreate, db: AsyncSession = Depends(get_db)):
    vo = VariationOrder(**data.model_dump())
    db.add(vo)
    await db.commit()
    await db.refresh(vo)
    return vo


@router.get("/projects/{project_id}/variation-orders")
async def list_vos(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(VariationOrder).where(VariationOrder.project_id == project_id))
    return result.scalars().all()


@router.get("/variation-orders/{vo_id}")
async def get_vo(vo_id: int, db: AsyncSession = Depends(get_db)):
    vo = await db.get(VariationOrder, vo_id)
    if not vo:
        raise HTTPException(404, "Variation Order not found")
    return vo


@router.put("/variation-orders/{vo_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_vo(vo_id: int, data: VariationOrderUpdate, db: AsyncSession = Depends(get_db)):
    vo = await db.get(VariationOrder, vo_id)
    if not vo:
        raise HTTPException(404, "Variation Order not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(vo, key, val)
    await db.commit()
    await db.refresh(vo)
    return vo


@router.delete("/variation-orders/{vo_id}", dependencies=[Depends(require_role("admin"))])
async def delete_vo(vo_id: int, db: AsyncSession = Depends(get_db)):
    vo = await db.get(VariationOrder, vo_id)
    if not vo:
        raise HTTPException(404, "Variation Order not found")
    await db.delete(vo)
    await db.commit()
    return {"detail": "Variation Order deleted successfully"}


# ─── VO BOQ Impact ───

@router.post("/variation-orders/{vo_id}/boq-items", dependencies=[Depends(require_role("admin", "engineer"))])
async def add_vo_boq_item(vo_id: int, data: VOBOQItemCreate, db: AsyncSession = Depends(get_db)):
    vo = await db.get(VariationOrder, vo_id)
    if not vo:
        raise HTTPException(404, "VO not found")
    boq = await db.get(BOQItem, data.boq_item_id)
    if not boq:
        raise HTTPException(404, "BOQ item not found")
    item_data = data.model_dump()
    item_data["vo_id"] = vo_id
    item = VOBOQItem(**item_data)
    db.add(item)
    vo.amount_change = (vo.amount_change or 0) + (data.quantity_change * data.unit_price_change)
    await db.commit()
    await db.refresh(item)
    return item


@router.get("/variation-orders/{vo_id}/boq-items")
async def list_vo_boq_items(vo_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(VOBOQItem).where(VOBOQItem.vo_id == vo_id))
    items = result.scalars().all()
    enriched = []
    for item in items:
        boq = await db.get(BOQItem, item.boq_item_id)
        enriched.append(VOBOQItemResponse(
            id=item.id, vo_id=item.vo_id, boq_item_id=item.boq_item_id,
            impact_type=item.impact_type, quantity_change=item.quantity_change,
            unit_price_change=item.unit_price_change, description=item.description,
            boq_item_code=boq.item_code if boq else None,
            boq_item_description=boq.description if boq else None,
        ))
    return enriched


@router.delete("/variation-orders/{vo_id}/boq-items/{item_id}", dependencies=[Depends(require_role("admin"))])
async def delete_vo_boq_item(vo_id: int, item_id: int, db: AsyncSession = Depends(get_db)):
    item = await db.get(VOBOQItem, item_id)
    if not item or item.vo_id != vo_id:
        raise HTTPException(404, "VO BOQ item not found")
    await db.delete(item)
    await db.commit()
    return {"detail": "VO BOQ item deleted"}


# ─── VO Schedule Impact ───

@router.post("/variation-orders/{vo_id}/schedule-impact", dependencies=[Depends(require_role("admin", "engineer"))])
async def add_vo_schedule_impact(vo_id: int, data: VOScheduleImpactCreate, db: AsyncSession = Depends(get_db)):
    vo = await db.get(VariationOrder, vo_id)
    if not vo:
        raise HTTPException(404, "VO not found")
    sched = await db.get(Schedule, data.schedule_id)
    if not sched:
        raise HTTPException(404, "Schedule not found")
    item_data = data.model_dump()
    item_data["vo_id"] = vo_id
    item = VOScheduleImpact(**item_data)
    db.add(item)
    vo.days_change = (vo.days_change or 0) + data.days_change
    await db.commit()
    await db.refresh(item)
    return item


@router.get("/variation-orders/{vo_id}/schedule-impacts")
async def list_vo_schedule_impacts(vo_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(VOScheduleImpact).where(VOScheduleImpact.vo_id == vo_id))
    items = result.scalars().all()
    enriched = []
    for item in items:
        sched = await db.get(Schedule, item.schedule_id)
        enriched.append(VOScheduleImpactResponse(
            id=item.id, vo_id=item.vo_id, schedule_id=item.schedule_id,
            days_change=item.days_change, description=item.description,
            task_name=sched.task_name if sched else None,
        ))
    return enriched


@router.delete("/variation-orders/{vo_id}/schedule-impacts/{impact_id}", dependencies=[Depends(require_role("admin"))])
async def delete_vo_schedule_impact(vo_id: int, impact_id: int, db: AsyncSession = Depends(get_db)):
    item = await db.get(VOScheduleImpact, impact_id)
    if not item or item.vo_id != vo_id:
        raise HTTPException(404, "VO schedule impact not found")
    await db.delete(item)
    await db.commit()
    return {"detail": "VO schedule impact deleted"}


# ─── VO Impact Summary ───

@router.get("/variation-orders/{vo_id}/impact-summary")
async def get_vo_impact_summary(vo_id: int, db: AsyncSession = Depends(get_db)):
    vo = await db.get(VariationOrder, vo_id)
    if not vo:
        raise HTTPException(404, "VO not found")
    boq_result = await db.execute(select(VOBOQItem).where(VOBOQItem.vo_id == vo_id))
    boq_items = boq_result.scalars().all()
    sched_result = await db.execute(select(VOScheduleImpact).where(VOScheduleImpact.vo_id == vo_id))
    sched_items = sched_result.scalars().all()
    boq_responses = []
    for item in boq_items:
        boq = await db.get(BOQItem, item.boq_item_id)
        boq_responses.append(VOBOQItemResponse(
            id=item.id, vo_id=item.vo_id, boq_item_id=item.boq_item_id,
            impact_type=item.impact_type, quantity_change=item.quantity_change,
            unit_price_change=item.unit_price_change, description=item.description,
            boq_item_code=boq.item_code if boq else None,
            boq_item_description=boq.description if boq else None,
        ))
    sched_responses = []
    for item in sched_items:
        sched = await db.get(Schedule, item.schedule_id)
        sched_responses.append(VOScheduleImpactResponse(
            id=item.id, vo_id=item.vo_id, schedule_id=item.schedule_id,
            days_change=item.days_change, description=item.description,
            task_name=sched.task_name if sched else None,
        ))
    return VOImpactSummary(
        vo_id=vo.id, vo_number=vo.vo_number, title=vo.title,
        total_amount_change=vo.amount_change, total_days_change=vo.days_change,
        approved_amount=vo.approved_amount, approved_days=vo.approved_days,
        affected_boq_items=boq_responses, affected_schedule_tasks=sched_responses,
    )


# ─── RFI ───

@router.post("/rfis", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_rfi(data: RFICreate, db: AsyncSession = Depends(get_db)):
    rfi = RFI(**data.model_dump())
    db.add(rfi)
    await db.commit()
    await db.refresh(rfi)
    return rfi


@router.get("/projects/{project_id}/rfis")
async def list_rfis(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(RFI).where(RFI.project_id == project_id))
    return result.scalars().all()


@router.get("/rfis/{rfi_id}")
async def get_rfi(rfi_id: int, db: AsyncSession = Depends(get_db)):
    rfi = await db.get(RFI, rfi_id)
    if not rfi:
        raise HTTPException(404, "RFI not found")
    return rfi


@router.put("/rfis/{rfi_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_rfi(rfi_id: int, data: RFIUpdate, db: AsyncSession = Depends(get_db)):
    rfi = await db.get(RFI, rfi_id)
    if not rfi:
        raise HTTPException(404, "RFI not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(rfi, key, val)
    await db.commit()
    await db.refresh(rfi)
    return rfi


@router.delete("/rfis/{rfi_id}", dependencies=[Depends(require_role("admin"))])
async def delete_rfi(rfi_id: int, db: AsyncSession = Depends(get_db)):
    rfi = await db.get(RFI, rfi_id)
    if not rfi:
        raise HTTPException(404, "RFI not found")
    await db.delete(rfi)
    await db.commit()
    return {"detail": "RFI deleted successfully"}


# ─── Material Approval Requests (MAR) ───

@router.post("/mar", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_mar(data: MARCreate, db: AsyncSession = Depends(get_db)):
    mar = MaterialApprovalRequest(**data.model_dump())
    db.add(mar)
    await db.commit()
    await db.refresh(mar)
    return mar


@router.get("/projects/{project_id}/mar")
async def list_project_mar(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(MaterialApprovalRequest)
        .where(MaterialApprovalRequest.project_id == project_id)
        .order_by(MaterialApprovalRequest.id.desc())
    )
    return result.scalars().all()


@router.get("/mar/{mar_id}")
async def get_mar(mar_id: int, db: AsyncSession = Depends(get_db)):
    mar = await db.get(MaterialApprovalRequest, mar_id)
    if not mar:
        raise HTTPException(404, "MAR not found")
    return mar


@router.put("/mar/{mar_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_mar(mar_id: int, data: MARUpdate, db: AsyncSession = Depends(get_db)):
    mar = await db.get(MaterialApprovalRequest, mar_id)
    if not mar:
        raise HTTPException(404, "MAR not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(mar, key, val)
    await db.commit()
    await db.refresh(mar)
    return mar


@router.post("/mar/{mar_id}/submit", dependencies=[Depends(require_role("admin", "engineer"))])
async def submit_mar(mar_id: int, db: AsyncSession = Depends(get_db)):
    mar = await db.get(MaterialApprovalRequest, mar_id)
    if not mar:
        raise HTTPException(404, "MAR not found")
    if mar.status != "draft":
        raise HTTPException(400, f"Cannot submit MAR in status '{mar.status}'")
    mar.status = "submitted"
    await db.commit()
    await db.refresh(mar)
    return mar


@router.post("/mar/{mar_id}/approve", dependencies=[Depends(require_role("admin", "engineer"))])
async def approve_mar(mar_id: int, db: AsyncSession = Depends(get_db)):
    mar = await db.get(MaterialApprovalRequest, mar_id)
    if not mar:
        raise HTTPException(404, "MAR not found")
    if mar.status not in ("submitted", "draft"):
        raise HTTPException(400, f"Cannot approve MAR in status '{mar.status}'")
    mar.status = "approved"
    await db.commit()
    await db.refresh(mar)
    return mar


@router.post("/mar/{mar_id}/reject", dependencies=[Depends(require_role("admin", "engineer"))])
async def reject_mar(mar_id: int, data: MARUpdate, db: AsyncSession = Depends(get_db)):
    mar = await db.get(MaterialApprovalRequest, mar_id)
    if not mar:
        raise HTTPException(404, "MAR not found")
    if mar.status not in ("submitted", "draft"):
        raise HTTPException(400, f"Cannot reject MAR in status '{mar.status}'")
    mar.status = "rejected"
    if data.rejection_reason:
        mar.rejection_reason = data.rejection_reason
    await db.commit()
    await db.refresh(mar)
    return mar


@router.delete("/mar/{mar_id}", dependencies=[Depends(require_role("admin"))])
async def delete_mar(mar_id: int, db: AsyncSession = Depends(get_db)):
    mar = await db.get(MaterialApprovalRequest, mar_id)
    if not mar:
        raise HTTPException(404, "MAR not found")
    await db.delete(mar)
    await db.commit()
    return {"detail": "MAR deleted"}


@router.post("/mar/{mar_id}/upload", dependencies=[Depends(require_role("admin", "engineer"))])
async def upload_mar_attachment(mar_id: int, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    mar = await db.get(MaterialApprovalRequest, mar_id)
    if not mar:
        raise HTTPException(404, "MAR not found")
    from app.core.config import settings
    import os
    upload_dir = os.path.join(settings.UPLOAD_DIR or "uploads", "mar")
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(file.filename)[1] if file.filename else ".pdf"
    fname = f"mar_{mar_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
    fpath = os.path.join(upload_dir, fname)
    content = await file.read()
    with open(fpath, "wb") as f:
        f.write(content)
    mar.file_path = fpath
    await db.commit()
    await db.refresh(mar)
    return mar


@router.get("/mar/{mar_id}/pdf")
async def export_mar_pdf(mar_id: int, db: AsyncSession = Depends(get_db)):
    mar = await db.get(MaterialApprovalRequest, mar_id)
    if not mar:
        raise HTTPException(404, "MAR not found")
    project = await db.get(Project, mar.project_id)
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate
    from reportlab.lib import colors
    import arabic_reshaper, bidi.algorithm
    from io import BytesIO

    def reshape(txt): return bidi.algorithm.get_display(arabic_reshaper.reshape(txt or ""))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>{reshape('طلب اعتماد مادة')}</b>", styles['Title']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>{reshape('رقم الطلب')}:</b> {reshape(mar.mar_number)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('العنوان')}:</b> {reshape(mar.title)}", styles['Normal']))
    if mar.specification_ref:
        story.append(Paragraph(f"<b>{reshape('مرجع المواصفات')}:</b> {reshape(mar.specification_ref)}", styles['Normal']))
    if mar.manufacturer:
        story.append(Paragraph(f"<b>{reshape('الشركة المصنعة')}:</b> {reshape(mar.manufacturer)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('النوع')}:</b> {reshape(mar.material_type)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('الكمية')}:</b> {mar.quantity_requested} {reshape(mar.unit)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('الحالة')}:</b> {reshape(mar.status)}", styles['Normal']))
    if project:
        story.append(Paragraph(f"<b>{reshape('المشروع')}:</b> {reshape(project.name)}", styles['Normal']))
    if mar.submitted_by:
        story.append(Paragraph(f"<b>{reshape('مقدم الطلب')}:</b> {reshape(mar.submitted_by)}", styles['Normal']))
    if mar.remarks:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('ملاحظات')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(mar.remarks), styles['Normal']))
    if mar.rejection_reason:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('سبب الرفض')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(mar.rejection_reason), styles['Normal']))

    story.append(Spacer(1, 20))
    status_text = {"draft": "مسودة", "submitted": "قيد المراجعة", "approved": "معتمد", "rejected": "مرفوض"}
    story.append(Paragraph(f"<b>{reshape('الحالة')}: {reshape(status_text.get(mar.status, mar.status))}</b>", styles['Normal']))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=mar_{mar.mar_number}.pdf"})


# ─── EVM (Earned Value Management) ───

@router.get("/projects/{project_id}/evm")
async def get_evm_report(project_id: int, db: AsyncSession = Depends(get_db)):
    ipcs = await db.execute(select(IPCHeader).where(IPCHeader.project_id == project_id))
    ipcs = ipcs.scalars().all()
    schedules = await db.execute(select(Schedule).where(Schedule.project_id == project_id))
    schedules = schedules.scalars().all()
    project = await db.get(Project, project_id)

    total_budget = project.budget_estimated or 0
    total_billed = sum(i.net_amount for i in ipcs if i.status in ("approved", "paid"))
    planned_progress = sum(s.progress_percent for s in schedules) / len(schedules) if schedules else 0

    planned_value = total_budget * (planned_progress / 100) if planned_progress else 0
    earned_value = total_billed
    actual_cost = total_billed * 0.95

    sv = earned_value - planned_value
    cv = earned_value - actual_cost
    spi = (earned_value / planned_value) if planned_value else 0
    cpi = (earned_value / actual_cost) if actual_cost else 0
    eac = (total_budget / cpi) if cpi else total_budget
    etc = eac - actual_cost
    vac = total_budget - eac

    return EVMReport(
        project_id=project_id,
        planned_value=Decimal(str(planned_value)),
        earned_value=Decimal(str(earned_value)),
        actual_cost=Decimal(str(actual_cost)),
        schedule_variance=Decimal(str(sv)),
        cost_variance=Decimal(str(cv)),
        spi=Decimal(str(spi)),
        cpi=Decimal(str(cpi)),
        estimate_at_completion=Decimal(str(eac)),
        estimate_to_complete=Decimal(str(etc)),
        variance_at_completion=Decimal(str(vac)),
        total_budget=total_budget,
        total_billed=total_billed,
    )


# ─── Reports ───

@router.get("/reports/project-financial/{project_id}")
async def report_project_financial(project_id: int, db: AsyncSession = Depends(get_db)):
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    contract_value = Decimal(str(project.budget_estimated or 0))

    boq_result = await db.execute(select(BOQItem).where(BOQItem.project_id == project_id))
    boq_items = boq_result.scalars().all()
    boq_total = sum((i.unit_price * i.quantity) for i in boq_items)

    ipc_result = await db.execute(select(IPCHeader).where(IPCHeader.project_id == project_id))
    ipcs = ipc_result.scalars().all()
    total_billed = sum(i.net_amount for i in ipcs if i.status in ("approved", "paid"))
    total_paid = sum(i.net_amount for i in ipcs if i.status == "paid")

    vo_result = await db.execute(select(VariationOrder).where(VariationOrder.project_id == project_id))
    vos = vo_result.scalars().all()
    total_vo_amount = sum(v.amount_change for v in vos)

    sched_result = await db.execute(select(Schedule).where(Schedule.project_id == project_id))
    schedules = sched_result.scalars().all()
    planned_progress = sum(s.progress_percent for s in schedules) / len(schedules) if schedules else Decimal("0")

    total_budget = contract_value
    planned_value = total_budget * (planned_progress / 100) if planned_progress else Decimal("0")
    earned_value = total_billed
    actual_cost = earned_value * Decimal("0.95")
    spi = (earned_value / planned_value) if planned_value else Decimal("0")
    cpi = (earned_value / actual_cost) if actual_cost else Decimal("0")
    remaining_budget = contract_value - total_paid
    percent_spent = (total_paid / contract_value * 100) if contract_value else Decimal("0")

    return ProjectFinancialReport(
        project_id=project_id,
        project_name=project.name,
        project_code=project.code,
        contract_value=contract_value,
        boq_total=boq_total,
        total_billed=total_billed,
        total_paid=total_paid,
        total_vo_amount=total_vo_amount,
        planned_value=Decimal(str(planned_value)),
        earned_value=earned_value,
        actual_cost=Decimal(str(actual_cost)),
        spi=Decimal(str(spi)),
        cpi=Decimal(str(cpi)),
        remaining_budget=remaining_budget,
        percent_spent=percent_spent,
    )


@router.get("/reports/project-comparison")
async def report_project_comparison(
    project_ids: str = Query(None, description="Comma-separated project IDs"),
    db: AsyncSession = Depends(get_db),
):
    if project_ids:
        ids = [int(x.strip()) for x in project_ids.split(",") if x.strip().isdigit()]
        result = await db.execute(select(Project).where(Project.id.in_(ids)))
    else:
        result = await db.execute(select(Project))
    projects = result.scalars().all()

    items = []
    for p in projects:
        ipc_result = await db.execute(select(IPCHeader).where(IPCHeader.project_id == p.id))
        ipcs = ipc_result.scalars().all()
        total_billed = sum(i.net_amount for i in ipcs if i.status in ("approved", "paid"))
        total_paid = sum(i.net_amount for i in ipcs if i.status == "paid")

        sched_result = await db.execute(select(Schedule).where(Schedule.project_id == p.id))
        schedules = sched_result.scalars().all()
        execution_rate = sum(s.progress_percent for s in schedules) / len(schedules) if schedules else Decimal("0")

        contract_value = Decimal(str(p.budget_estimated or 0))
        financial_progress = (total_paid / contract_value * 100) if contract_value else Decimal("0")

        items.append(ProjectComparisonItem(
            id=p.id,
            code=p.code,
            name=p.name,
            contract_value=contract_value,
            total_billed=total_billed,
            total_paid=total_paid,
            execution_rate=Decimal(str(execution_rate)),
            financial_progress=Decimal(str(financial_progress)),
        ))

    return items


@router.get("/reports/dashboard-export")
async def report_dashboard_export(db: AsyncSession = Depends(get_db)):
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(["=== Projects ==="])
    writer.writerow(["ID", "Code", "Name", "Status", "Contract Value", "Location", "Client"])
    projects = (await db.execute(select(Project))).scalars().all()
    for p in projects:
        writer.writerow([p.id, p.code, p.name, p.status, float(p.budget_estimated or 0), p.location, p.client_name])

    writer.writerow([])
    writer.writerow(["=== Contracts ==="])
    writer.writerow(["ID", "Project ID", "Contract Number", "Type", "Value", "Status"])
    contracts = (await db.execute(select(Contract))).scalars().all()
    for c in contracts:
        writer.writerow([c.id, c.project_id, c.contract_number, c.contract_type, float(c.value), c.status])

    writer.writerow([])
    writer.writerow(["=== BOQ Items ==="])
    writer.writerow(["ID", "Project ID", "Item Code", "Description", "Unit", "Quantity", "Unit Price", "Total Price"])
    boq_items = (await db.execute(select(BOQItem))).scalars().all()
    for b in boq_items:
        writer.writerow([b.id, b.project_id, b.item_code, b.description, b.unit, float(b.quantity), float(b.unit_price), float(b.total_price)])

    writer.writerow([])
    writer.writerow(["=== IPCs ==="])
    writer.writerow(["ID", "Project ID", "IPC Number", "Period", "Status", "Net Amount", "Total Works", "Gross Amount"])
    ipcs = (await db.execute(select(IPCHeader))).scalars().all()
    for i in ipcs:
        writer.writerow([i.id, i.project_id, i.ipc_number, i.ipc_period, i.status, float(i.net_amount), float(i.total_works), float(i.gross_amount)])

    writer.writerow([])
    writer.writerow(["=== Variation Orders ==="])
    writer.writerow(["ID", "Project ID", "VO Number", "Title", "Amount Change", "Status"])
    vos = (await db.execute(select(VariationOrder))).scalars().all()
    for v in vos:
        writer.writerow([v.id, v.project_id, v.vo_number, v.title, float(v.amount_change), v.status])

    writer.writerow([])
    writer.writerow(["=== Schedules ==="])
    writer.writerow(["ID", "Project ID", "Task Name", "Duration Days", "Progress %", "Status"])
    schedules = (await db.execute(select(Schedule))).scalars().all()
    for s in schedules:
        writer.writerow([s.id, s.project_id, s.task_name, s.duration_days, float(s.progress_percent), s.status])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=dashboard_export.csv"},
    )


# ─── Non-Conformance Reports (NCR) ───

@router.post("/ncr", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_ncr(data: NCRCreate, db: AsyncSession = Depends(get_db)):
    ncr = NonConformanceReport(**data.model_dump())
    db.add(ncr)
    await db.commit()
    await db.refresh(ncr)
    return ncr


@router.get("/projects/{project_id}/ncr")
async def list_project_ncr(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(NonConformanceReport)
        .where(NonConformanceReport.project_id == project_id)
        .order_by(NonConformanceReport.id.desc())
    )
    return result.scalars().all()


@router.get("/ncr/{ncr_id}")
async def get_ncr(ncr_id: int, db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    return ncr


@router.put("/ncr/{ncr_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_ncr(ncr_id: int, data: NCRUpdate, db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(ncr, key, val)
    await db.commit()
    await db.refresh(ncr)
    return ncr


@router.post("/ncr/{ncr_id}/investigate", dependencies=[Depends(require_role("admin", "engineer"))])
async def investigate_ncr(ncr_id: int, db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    if ncr.status != "open":
        raise HTTPException(400, f"Cannot investigate NCR in status '{ncr.status}'")
    ncr.status = "investigation"
    await db.commit()
    await db.refresh(ncr)
    return ncr


@router.post("/ncr/{ncr_id}/action", dependencies=[Depends(require_role("admin", "engineer"))])
async def action_ncr(ncr_id: int, data: NCRUpdate, db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    if ncr.status != "investigation":
        raise HTTPException(400, f"Cannot apply corrective action in status '{ncr.status}'")
    ncr.status = "corrective_action"
    if data.corrective_action:
        ncr.corrective_action = data.corrective_action
    if data.preventive_action:
        ncr.preventive_action = data.preventive_action
    await db.commit()
    await db.refresh(ncr)
    return ncr


@router.post("/ncr/{ncr_id}/close", dependencies=[Depends(require_role("admin", "engineer"))])
async def close_ncr(ncr_id: int, data: NCRUpdate, db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    if ncr.status not in ("corrective_action", "investigation"):
        raise HTTPException(400, f"Cannot close NCR in status '{ncr.status}'")
    ncr.status = "closed"
    ncr.closed_date = data.closed_date or date.today()
    await db.commit()
    await db.refresh(ncr)
    return ncr


@router.post("/ncr/{ncr_id}/reject", dependencies=[Depends(require_role("admin"))])
async def reject_ncr(ncr_id: int, data: NCRUpdate, db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    ncr.status = "rejected"
    if data.rejection_reason:
        ncr.rejection_reason = data.rejection_reason
    await db.commit()
    await db.refresh(ncr)
    return ncr


@router.delete("/ncr/{ncr_id}", dependencies=[Depends(require_role("admin"))])
async def delete_ncr(ncr_id: int, db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    await db.delete(ncr)
    await db.commit()
    return {"detail": "NCR deleted"}


# ─── NCR Attachments ───

@router.post("/ncr/{ncr_id}/upload", dependencies=[Depends(require_role("admin", "engineer"))])
async def upload_ncr_attachment(ncr_id: int, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    import os
    upload_dir = os.path.join("uploads", "ncr")
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(file.filename)[1] if file.filename else ".jpg"
    fname = f"ncr_{ncr_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
    fpath = os.path.join(upload_dir, fname)
    content = await file.read()
    with open(fpath, "wb") as f:
        f.write(content)
    ncr.file_path = fpath
    await db.commit()
    await db.refresh(ncr)
    return ncr


# ─── NCR PDF Report ───

@router.get("/ncr/{ncr_id}/pdf")
async def export_ncr_pdf(ncr_id: int, db: AsyncSession = Depends(get_db)):
    ncr = await db.get(NonConformanceReport, ncr_id)
    if not ncr:
        raise HTTPException(404, "NCR not found")
    project = await db.get(Project, ncr.project_id)
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, Spacer, SimpleDocTemplate
    import arabic_reshaper, bidi.algorithm
    from io import BytesIO

    def reshape(txt): return bidi.algorithm.get_display(arabic_reshaper.reshape(txt or ""))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>{reshape('تقرير عدم المطابقة')}</b>", styles['Title']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>{reshape('رقم التقرير')}:</b> {reshape(ncr.ncr_number)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('العنوان')}:</b> {reshape(ncr.title)}", styles['Normal']))
    if ncr.description:
        story.append(Paragraph(f"<b>{reshape('الوصف')}:</b> {reshape(ncr.description)}", styles['Normal']))
    if ncr.location:
        story.append(Paragraph(f"<b>{reshape('الموقع')}:</b> {reshape(ncr.location)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('التصنيف')}:</b> {reshape(ncr.category)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('الخطورة')}:</b> {reshape(ncr.severity)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('المصدر')}:</b> {reshape(ncr.source)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('الحالة')}:</b> {reshape(ncr.status)}", styles['Normal']))
    if project:
        story.append(Paragraph(f"<b>{reshape('المشروع')}:</b> {reshape(project.name)}", styles['Normal']))
    if ncr.identified_by:
        story.append(Paragraph(f"<b>{reshape('مكتشف')}:</b> {reshape(ncr.identified_by)}", styles['Normal']))
    if ncr.corrective_action:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('الإجراء التصحيحي')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(ncr.corrective_action), styles['Normal']))
    if ncr.preventive_action:
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"<b>{reshape('الإجراء الوقائي')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(ncr.preventive_action), styles['Normal']))
    if ncr.rejection_reason:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('سبب الرفض')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(ncr.rejection_reason), styles['Normal']))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=ncr_{ncr.ncr_number}.pdf"})


# ─── Meeting Minutes ───

@router.post("/meeting-minutes", dependencies=[Depends(require_role("admin", "engineer"))])
async def create_meeting_minute(data: MeetingMinuteCreate, db: AsyncSession = Depends(get_db)):
    mm = MeetingMinute(**data.model_dump())
    db.add(mm)
    await db.commit()
    await db.refresh(mm)
    return mm


@router.get("/projects/{project_id}/meeting-minutes")
async def list_meeting_minutes(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(MeetingMinute)
        .where(MeetingMinute.project_id == project_id)
        .order_by(MeetingMinute.meeting_date.desc())
    )
    return result.scalars().all()


@router.get("/meeting-minutes/{mm_id}")
async def get_meeting_minute(mm_id: int, db: AsyncSession = Depends(get_db)):
    mm = await db.get(MeetingMinute, mm_id)
    if not mm:
        raise HTTPException(404, "Meeting minute not found")
    return mm


@router.put("/meeting-minutes/{mm_id}", dependencies=[Depends(require_role("admin", "engineer"))])
async def update_meeting_minute(mm_id: int, data: MeetingMinuteUpdate, db: AsyncSession = Depends(get_db)):
    mm = await db.get(MeetingMinute, mm_id)
    if not mm:
        raise HTTPException(404, "Meeting minute not found")
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(mm, key, val)
    await db.commit()
    await db.refresh(mm)
    return mm


@router.post("/meeting-minutes/{mm_id}/finalize", dependencies=[Depends(require_role("admin", "engineer"))])
async def finalize_meeting_minute(mm_id: int, db: AsyncSession = Depends(get_db)):
    mm = await db.get(MeetingMinute, mm_id)
    if not mm:
        raise HTTPException(404, "Meeting minute not found")
    if mm.status != "draft":
        raise HTTPException(400, f"Cannot finalize in status '{mm.status}'")
    mm.status = "final"
    await db.commit()
    await db.refresh(mm)
    return mm


@router.delete("/meeting-minutes/{mm_id}", dependencies=[Depends(require_role("admin"))])
async def delete_meeting_minute(mm_id: int, db: AsyncSession = Depends(get_db)):
    mm = await db.get(MeetingMinute, mm_id)
    if not mm:
        raise HTTPException(404, "Meeting minute not found")
    await db.delete(mm)
    await db.commit()
    return {"detail": "Meeting minute deleted"}


@router.get("/meeting-minutes/{mm_id}/pdf")
async def export_meeting_minute_pdf(mm_id: int, db: AsyncSession = Depends(get_db)):
    mm = await db.get(MeetingMinute, mm_id)
    if not mm:
        raise HTTPException(404, "Meeting minute not found")
    project = await db.get(Project, mm.project_id)
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, Spacer, SimpleDocTemplate
    import arabic_reshaper, bidi.algorithm
    from io import BytesIO

    def reshape(txt):
        return bidi.algorithm.get_display(arabic_reshaper.reshape(txt or ""))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>{reshape('محضر اجتماع')}</b>", styles['Title']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>{reshape('عنوان الاجتماع')}:</b> {reshape(mm.meeting_title)}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('التاريخ')}:</b> {mm.meeting_date.isoformat() if mm.meeting_date else ''}", styles['Normal']))
    story.append(Paragraph(f"<b>{reshape('النوع')}:</b> {reshape(mm.meeting_type)}", styles['Normal']))
    if mm.location:
        story.append(Paragraph(f"<b>{reshape('الموقع')}:</b> {reshape(mm.location)}", styles['Normal']))
    if mm.chairperson:
        story.append(Paragraph(f"<b>{reshape('رئيس الاجتماع')}:</b> {reshape(mm.chairperson)}", styles['Normal']))
    if project:
        story.append(Paragraph(f"<b>{reshape('المشروع')}:</b> {reshape(project.name)}", styles['Normal']))

    if mm.attendees:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('الحضور')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(mm.attendees), styles['Normal']))

    if mm.agenda:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('جدول الأعمال')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(mm.agenda), styles['Normal']))

    if mm.discussion:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('المناقشات')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(mm.discussion), styles['Normal']))

    if mm.decisions:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('القرارات')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(mm.decisions), styles['Normal']))

    if mm.action_items:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('بنود العمل')}:</b>", styles['Normal']))
        story.append(Paragraph(reshape(mm.action_items), styles['Normal']))

    if mm.next_meeting_date:
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>{reshape('الاجتماع القادم')}:</b> {mm.next_meeting_date.isoformat()}", styles['Normal']))

    story.append(Spacer(1, 20))
    status_text = {"draft": "مسودة", "final": "نهائي"}
    story.append(Paragraph(f"<b>{reshape('الحالة')}: {reshape(status_text.get(mm.status, mm.status))}</b>", styles['Normal']))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=meeting_minutes_{mm_id}.pdf"})


# ─── Notification Helper ───


async def _create_notification(db: AsyncSession, user_id: int, title: str, message: str = None, type: str = "info", link: str = None):
    notif = Notification(user_id=user_id, title=title, message=message, type=type, link=link)
    db.add(notif)
    await db.flush()
    return notif


# ─── Notifications ───


@router.get("/notifications")
async def list_notifications(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    result = await db.execute(
        select(Notification)
        .where(Notification.user_id == user.id)
        .order_by(Notification.created_at.desc())
        .offset(skip).limit(limit)
    )
    items = result.scalars().all()
    return {"items": items, "total": len(items)}


@router.get("/notifications/unread-count")
async def unread_notification_count(
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    result = await db.execute(
        select(func.count(Notification.id)).where(
            Notification.user_id == user.id,
            Notification.is_read == False,
        )
    )
    return {"count": result.scalar() or 0}


@router.put("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: int,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    notif = await db.get(Notification, notification_id)
    if not notif or notif.user_id != user.id:
        raise HTTPException(404, "Notification not found")
    notif.is_read = True
    await db.commit()
    return {"detail": "Notification marked as read"}


@router.put("/notifications/read-all")
async def mark_all_notifications_read(
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    await db.execute(
        update(Notification).where(Notification.user_id == user.id, Notification.is_read == False).values(is_read=True)
    )
    await db.commit()
    return {"detail": "All notifications marked as read"}


# ─── Admin: System Logs ───

@router.get("/admin/logs", dependencies=[Depends(require_role("admin"))])
async def get_system_logs(lines: int = Query(100, ge=1, le=1000)):
    import os
    log_paths = [
        os.path.join(os.path.dirname(os.path.dirname(__file__)), "backend.log"),
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "backend.log"),
        "backend.log",
    ]
    log_file = None
    for p in log_paths:
        if os.path.exists(p):
            log_file = p
            break
    if not log_file:
        return {"lines": []}
    with open(log_file, "r", encoding="utf-8", errors="ignore") as f:
        all_lines = f.readlines()
    last_lines = all_lines[-lines:]
    return {"lines": [line.rstrip("\n\r") for line in last_lines]}


# ─── Admin: System Settings ───

@router.get("/admin/settings", dependencies=[Depends(require_role("admin"))])
async def list_settings(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SystemSetting).order_by(SystemSetting.key))
    return result.scalars().all()


@router.post("/admin/settings", dependencies=[Depends(require_role("admin"))])
async def create_setting(data: SystemSettingCreate, db: AsyncSession = Depends(get_db)):
    existing = await db.execute(select(SystemSetting).where(SystemSetting.key == data.key))
    if existing.scalar_one_or_none():
        raise HTTPException(400, "Setting key already exists")
    setting = SystemSetting(key=data.key, value=data.value, description=data.description)
    db.add(setting)
    await db.commit()
    await db.refresh(setting)
    return setting


@router.put("/admin/settings/{key}", dependencies=[Depends(require_role("admin"))])
async def update_setting(key: str, data: SystemSettingUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SystemSetting).where(SystemSetting.key == key))
    setting = result.scalar_one_or_none()
    if not setting:
        raise HTTPException(404, "Setting not found")
    setting.value = data.value
    await db.commit()
    await db.refresh(setting)
    return setting


# ─── Admin: Activity Log ───

@router.get("/admin/activity", dependencies=[Depends(require_role("admin"))])
async def list_activity(
    limit: int = Query(50, ge=1, le=500),
    action: str = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(AuditLog).order_by(AuditLog.created_at.desc()).limit(limit)
    if action:
        stmt = stmt.where(AuditLog.action == action)
    result = await db.execute(stmt)
    logs = result.scalars().all()

    user_ids = list(set(log.user_id for log in logs))
    users_result = await db.execute(select(User).where(User.id.in_(user_ids)))
    users_map = {u.id: u.username for u in users_result.scalars().all()}

    return [
        ActivityLogResponse(
            id=log.id,
            user_id=log.user_id,
            username=users_map.get(log.user_id, "Unknown"),
            action=log.action,
            resource=log.entity_type,
            details=log.details,
            ip_address=None,
            created_at=log.created_at,
        )
        for log in logs
    ]
